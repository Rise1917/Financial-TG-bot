"""Парсер CSV и Excel выписок (Kaspi, Halyk, конвертированные PDF и др.)."""

import csv
import io
import logging
import re
from typing import Any

from openpyxl import load_workbook

from bot.services.statements.models import ParsedTransaction, ParseResult
from bot.services.statements.parsers.kaspi import extract_transactions_from_grid
from bot.services.statements.parsers.utils import detect_period, parse_amount, parse_date

logger = logging.getLogger(__name__)

_DATE_ALIASES = ("дата", "date", "күні", "дата операции", "transaction date", "время")
_DESC_ALIASES = (
    "описание", "назначение", "детали", "description", "merchant",
    "получатель", "контрагент", "магазин", "назначение платежа", "операция",
)
_AMOUNT_ALIASES = ("сумма", "amount", "сома", "итого", "сумма операции")
_DEBIT_ALIASES = ("дебет", "списание", "расход", "debit", "withdrawal", "снятие")
_CREDIT_ALIASES = ("кредит", "поступление", "приход", "credit", "deposit", "зачисление")
_CATEGORY_ALIASES = ("категория", "category", "тип", "вид операции")


def parse_csv(content: bytes, bank: str = "") -> ParseResult:
    text = _decode_bytes(content)
    delimiter = ";" if text.count(";") > text.count(",") else ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    rows = list(reader)

    if not rows:
        return ParseResult(errors=["CSV-файл пуст или без заголовков"], bank_detected=bank)

    structured = _parse_rows(rows, bank or "CSV")
    if structured.transactions:
        return structured

    grid = [tuple(row.values()) for row in rows]
    return _parse_grid(grid, bank or "CSV")


def parse_excel(content: bytes, bank: str = "") -> ParseResult:
    bank_label = bank or "Excel"

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        return ParseResult(errors=[f"Не удалось открыть Excel: {exc}"], bank_detected=bank_label)

    if not workbook.sheetnames:
        return ParseResult(errors=["Excel-файл не содержит листов"], bank_detected=bank_label)

    sheet_count = len(workbook.sheetnames)
    all_heuristic: list[ParsedTransaction] = []
    all_structured: list[ParsedTransaction] = []
    seen: set[tuple[str, float, str]] = set()

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        raw_rows = list(sheet.iter_rows(values_only=True))
        if not raw_rows:
            continue

        logger.info(
            "Excel лист '%s': строк=%d",
            sheet_name,
            len(raw_rows),
        )

        heuristic = extract_transactions_from_grid(raw_rows, bank_label)
        for tx in heuristic:
            key = (tx.date[:10], tx.amount, tx.description[:60])
            if key not in seen:
                seen.add(key)
                all_heuristic.append(tx)

        structured_rows = _rows_from_header_table(raw_rows)
        if structured_rows:
            structured = _parse_rows(structured_rows, bank_label)
            for tx in structured.transactions:
                key = (tx.date[:10], tx.amount, tx.description[:60])
                if key not in seen:
                    seen.add(key)
                    all_structured.append(tx)

    transactions = all_heuristic if len(all_heuristic) >= len(all_structured) else all_structured
    if len(all_heuristic) > 0 and len(all_structured) > 0:
        combined_seen: set[tuple[str, float, str]] = set()
        transactions = []
        for tx in all_heuristic + all_structured:
            key = (tx.date[:10], tx.amount, tx.description[:60])
            if key not in combined_seen:
                combined_seen.add(key)
                transactions.append(tx)

    dates = [tx.date for tx in transactions]
    result = ParseResult(
        transactions=transactions,
        bank_detected=bank_label,
        period_from=detect_period(dates)[0],
        period_to=detect_period(dates)[1],
    )

    if not transactions:
        result.errors.append(
            "Не удалось распознать операции в Excel. "
            "Если файл получен конвертацией из PDF — отправьте оригинальный PDF, "
            "бот умеет разбирать его напрямую."
        )
    else:
        logger.info(
            "Excel разобран: листов=%d, операций=%d",
            sheet_count,
            len(transactions),
        )

    workbook.close()
    return result


def _parse_grid(grid: list[tuple[Any, ...]], bank: str) -> ParseResult:
    transactions = extract_transactions_from_grid(grid, bank)
    dates = [tx.date for tx in transactions]
    result = ParseResult(
        transactions=transactions,
        bank_detected=bank,
        period_from=detect_period(dates)[0],
        period_to=detect_period(dates)[1],
    )
    if not transactions:
        result.errors.append("Не удалось распознать операции в таблице")
    return result


def _rows_from_header_table(raw_rows: list[tuple[Any, ...]]) -> list[dict[str, Any]]:
    header_idx = _find_header_row(raw_rows, max_scan=60)
    if header_idx is None:
        return []

    headers = [_normalize_header(cell) for cell in raw_rows[header_idx]]
    rows: list[dict[str, Any]] = []

    for raw in raw_rows[header_idx + 1 :]:
        if not any(cell is not None and str(cell).strip() for cell in raw):
            continue
        row = {
            headers[i]: raw[i] if i < len(raw) else None
            for i in range(len(headers))
            if headers[i]
        }
        rows.append(row)

    return rows


def _parse_rows(rows: list[dict[str, Any]], bank: str) -> ParseResult:
    result = ParseResult(bank_detected=bank)
    dates: list[str] = []

    if not rows:
        result.errors.append("Таблица пуста")
        return result

    col_map = _map_columns(rows[0].keys())

    for row in rows:
        date_str = parse_date(_get_cell(row, col_map.get("date", "")))
        description = str(_get_cell(row, col_map.get("desc", "")) or "").strip()
        bank_category = str(_get_cell(row, col_map.get("category", "")) or "").strip()

        amount = None
        is_expense = True

        if col_map.get("debit") or col_map.get("credit"):
            debit = parse_amount(_get_cell(row, col_map.get("debit", "")))
            credit = parse_amount(_get_cell(row, col_map.get("credit", "")))
            if debit:
                amount, is_expense = debit, True
            elif credit:
                amount, is_expense = credit, False
        else:
            raw_amount = _get_cell(row, col_map.get("amount", ""))
            amount = parse_amount(raw_amount)
            if amount and isinstance(raw_amount, str) and raw_amount.strip().startswith(("-", "−")):
                is_expense = True
            elif amount and bank_category:
                cat_lower = bank_category.lower()
                is_expense = not any(
                    word in cat_lower for word in ("поступ", "пополн", "зачисл", "возврат")
                )

        if not date_str or not amount:
            continue

        if not description:
            description = bank_category or "Операция по выписке"

        dates.append(date_str)
        result.transactions.append(
            ParsedTransaction(
                date=date_str,
                amount=amount,
                is_expense=is_expense,
                description=description[:500],
                bank_category=bank_category,
            )
        )

    result.period_from, result.period_to = detect_period(dates)
    if not result.transactions:
        result.errors.append("Не удалось распознать операции в таблице")
    return result


def _decode_bytes(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1251", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="ignore")


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip().lower())


def _find_header_row(
    rows: list[tuple[Any, ...]], max_scan: int = 60
) -> int | None:
    for idx, row in enumerate(rows[:max_scan]):
        headers = [_normalize_header(cell) for cell in row]
        joined = " ".join(headers)
        if any(alias in joined for alias in _DATE_ALIASES) and (
            any(alias in joined for alias in _AMOUNT_ALIASES)
            or any(alias in joined for alias in _DEBIT_ALIASES)
        ):
            return idx
    return None


def _map_columns(columns: Any) -> dict[str, str]:
    normalized = {_normalize_header(col): col for col in columns}
    mapping: dict[str, str] = {}

    for key, original in normalized.items():
        if _matches(key, _DATE_ALIASES):
            mapping["date"] = original
        elif _matches(key, _DESC_ALIASES):
            mapping["desc"] = original
        elif _matches(key, _DEBIT_ALIASES):
            mapping["debit"] = original
        elif _matches(key, _CREDIT_ALIASES):
            mapping["credit"] = original
        elif _matches(key, _CATEGORY_ALIASES):
            mapping["category"] = original
        elif _matches(key, _AMOUNT_ALIASES) and "amount" not in mapping:
            mapping["amount"] = original

    return mapping


def _matches(header: str, aliases: tuple[str, ...]) -> bool:
    return any(alias in header for alias in aliases)


def _get_cell(row: dict[str, Any], column: str) -> Any:
    if not column:
        return None
    return row.get(column)
